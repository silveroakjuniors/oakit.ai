from fastapi import FastAPI, BackgroundTasks, UploadFile, File, Form, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from dotenv import load_dotenv
from datetime import datetime, date
import hashlib
import io

load_dotenv()

app = FastAPI(title="Oakit AI Service", version="0.1.0")

THOUGHTS_FOR_DAY = [
    "A great teacher doesn't just teach — they inspire curiosity.",
    "Every child learns differently. Patience is your greatest tool.",
    "Small moments of encouragement can change a student's entire day.",
    "The best classroom is one where every student feels seen.",
    "Teaching is the one profession that creates all other professions.",
    "Your energy sets the tone for the entire classroom.",
    "A smile and a kind word can unlock a child's potential.",
    "Celebrate small wins — they build the confidence for big ones.",
    "The questions children ask are the seeds of tomorrow's discoveries.",
    "You are not just teaching a subject — you are shaping a future.",
    "Consistency and care are the foundations of great teaching.",
    "Every lesson is a chance to make learning feel like an adventure.",
    "The most powerful thing you can give a student is belief in themselves.",
    "Great teachers listen as much as they speak.",
    "Your dedication today plants seeds that bloom for a lifetime.",
]


@app.get("/health")
async def health():
    return {"status": "ok", "service": "oakit-ai-service"}


# --- Plan-scoped questions generator ---

class PlanQuestionsRequest(BaseModel):
    teacher_id: str
    school_id: str
    query_date: str

@app.post("/internal/plan-questions")
async def plan_questions(req: PlanQuestionsRequest):
    """Generate suggested questions based on today's plan — strictly scoped to curriculum."""
    from query_pipeline import generate_plan_questions
    result = await generate_plan_questions(req.teacher_id, req.school_id, req.query_date)
    return result


# --- Curriculum preview (Week 1 only, before full ingestion) ---

@app.post("/internal/preview")
async def preview_curriculum(file: UploadFile = File(...), start_page: str = "1"):
    """Extract only Week 1 from the PDF and return it for admin approval."""
    import tempfile, os
    from extractor import extract_days, extract_pages
    from chunker import chunk_from_days

    # start_page comes as form string
    try:
        start_page_int = int(start_page)
    except (ValueError, TypeError):
        start_page_int = 1

    content = await file.read()
    suffix = os.path.splitext(file.filename or "upload.pdf")[1] or ".pdf"

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        day_entries, failed_pages = extract_days(tmp_path, start_page=start_page_int)
        # Filter to Week 1 only
        week1 = [d for d in day_entries if d.week == 1]

        if not week1:
            # Fallback: return first 5 text chunks
            pages, _ = extract_pages(tmp_path, start_page=start_page_int)
            from chunker import chunk_document
            chunks = chunk_document(pages)[:5]
            return {
                "mode": "text",
                "week": 1,
                "total_weeks_detected": 0,
                "total_days_detected": len(day_entries),
                "days": [{"week": 0, "day": i+1, "topic_label": c.topic_label, "subjects": {"Content": c.content[:500]}} for i, c in enumerate(chunks)],
                "failed_pages": failed_pages,
            }

        total_weeks = max((d.week for d in day_entries), default=0)
        total_days = len(day_entries)

        return {
            "mode": "table",
            "week": 1,
            "total_weeks_detected": total_weeks,
            "total_days_detected": total_days,
            "days": [
                {
                    "week": d.week,
                    "day": d.day,
                    "topic_label": d.topic_label,
                    "subjects": d.subjects,
                }
                for d in week1
            ],
            "failed_pages": failed_pages,
        }
    except Exception as e:
        import traceback
        print(f"Preview error: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


# --- Ingestion ---

class IngestRequest(BaseModel):
    document_id: str

@app.post("/internal/ingest")
async def ingest(req: IngestRequest, background_tasks: BackgroundTasks):
    background_tasks.add_task(_run_ingest, req.document_id)
    return {"message": "Ingestion started", "document_id": req.document_id}

async def _run_ingest(document_id: str):
    try:
        from ingestion_service import ingest_document
        result = await ingest_document(document_id)
        print(f"Ingestion complete for {document_id}: {result}")
    except Exception as e:
        print(f"Ingestion failed for {document_id}: {e}")


# --- Query ---

class QueryRequest(BaseModel):
    teacher_id: str
    school_id: str
    text: str
    query_date: str
    role: str = "teacher"
    history: list = []
    context: str = ""  # pre-built context for parent queries

@app.post("/internal/query")
async def query_endpoint(req: QueryRequest):
    from query_pipeline import query
    result = await query(req.teacher_id, req.school_id, req.text, req.query_date, req.role, history=req.history, context=req.context)
    return result


# ─── Session transcript formatter ────────────────────────────────────────────

class FormatSessionRequest(BaseModel):
    raw_transcript: str
    class_context: str = ""

@app.post("/internal/format-session")
async def format_session(req: FormatSessionRequest):
    """
    Format a raw speech transcript into clean class notes using Gemini.
    Falls back to structured plain text if Gemini is unavailable.
    """
    from query_pipeline import _call_llm

    prompt = f"""You are a school assistant helping a teacher format their classroom session notes.

{f"Context: {req.class_context}" if req.class_context else ""}

The teacher recorded this session transcript using voice recognition:
---
{req.raw_transcript[:8000]}
---

Please format this into clean, structured class notes that can be shared with parents. Include:
1. A brief summary of what was covered today (2-3 sentences)
2. Key topics discussed (bullet points)
3. Any activities or exercises mentioned
4. Homework or follow-up items if mentioned

Keep it concise, clear, and parent-friendly. Use simple language.
Format with clear sections using emoji headers like 📚 Topics Covered, 🎯 Activities, 📝 Homework."""

    system = "You are a helpful school assistant. Format classroom notes clearly and concisely for parents."

    formatted, _ = await _call_llm(prompt, system)

    if not formatted:
        # Fallback: basic structure without AI
        lines = req.raw_transcript.strip().split('. ')
        formatted = f"📚 Class Session Notes\n\n"
        formatted += f"📝 Session Summary:\n{req.raw_transcript[:500]}{'...' if len(req.raw_transcript) > 500 else ''}\n\n"
        formatted += f"ℹ️ Note: This is the raw transcript from today's session."

    return {"formatted": formatted}


# ─── Voice transcription ──────────────────────────────────────────────────────

@app.post("/internal/transcribe")
async def transcribe_audio(
    audio: UploadFile = File(...),
    language: str = Form(default="en"),
):
    """
    Transcribe audio to text using Gemini's audio understanding.
    Accepts any audio format (webm, mp4, wav, ogg, m4a).
    Returns: { transcript: str, language: str, duration_seconds: float }
    """
    import os, httpx, base64, mimetypes

    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(status_code=503, detail="Voice transcription not configured")

    # Read audio bytes
    audio_bytes = await audio.read()
    if len(audio_bytes) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=413, detail="Audio file too large (max 10MB)")
    if len(audio_bytes) < 100:
        raise HTTPException(status_code=400, detail="Audio file too small or empty")

    # Determine MIME type
    content_type = audio.content_type or "audio/webm"
    # Normalize common types
    mime_map = {
        "audio/webm": "audio/webm",
        "audio/ogg": "audio/ogg",
        "audio/wav": "audio/wav",
        "audio/mp4": "audio/mp4",
        "audio/m4a": "audio/mp4",
        "audio/mpeg": "audio/mpeg",
        "video/webm": "audio/webm",  # Chrome records as video/webm
    }
    mime_type = mime_map.get(content_type, "audio/webm")

    audio_b64 = base64.b64encode(audio_bytes).decode("utf-8")

    # Language hint for the prompt
    lang_hints = {
        "hi": "Hindi", "te": "Telugu", "kn": "Kannada", "ta": "Tamil",
        "ml": "Malayalam", "gu": "Gujarati", "mr": "Marathi", "bn": "Bengali",
        "pa": "Punjabi", "ur": "Urdu", "ar": "Arabic", "fr": "French",
        "es": "Spanish", "en": "English",
    }
    lang_name = lang_hints.get(language, "English")

    prompt = (
        f"Transcribe this audio recording accurately. "
        f"The speaker is likely speaking in {lang_name}. "
        f"This is a school-related question from a teacher or parent. "
        f"Return ONLY the transcribed text, nothing else. "
        f"If the audio is unclear or silent, return an empty string."
    )

    try:
        # ── PHASE 2: Real Gemini transcription (not yet available on this API key)
        # For now, return a mock transcript to demonstrate the voice flow in demos.
        # The audio IS recorded and received correctly — only the transcription step is mocked.
        print(f"[Transcribe] DEMO MODE — received {len(audio_bytes)} bytes of {mime_type} audio, returning mock transcript")

        # Pick a contextual demo response based on language
        demo_transcripts = {
            "hi": "आज का पाठ्यक्रम क्या है?",
            "te": "ఈరోజు పాఠ్యప్రణాళిక ఏమిటి?",
            "kn": "ಇಂದಿನ ಪಾಠ್ಯಕ್ರಮ ಏನು?",
            "ta": "இன்றைய பாடத்திட்டம் என்ன?",
            "en": "What is today's plan for my child?",
        }
        mock_transcript = demo_transcripts.get(language, demo_transcripts["en"])
        return {"transcript": mock_transcript, "language": language, "demo_mode": True}

    except HTTPException:
        raise
    except Exception as e:
        print(f"[Transcribe] Exception: {e}")
        raise HTTPException(status_code=500, detail=f"Transcription failed: {str(e)}")




class PlanRequest(BaseModel):
    class_id: str
    section_id: str
    school_id: str
    academic_year: str
    month: int = None
    plan_year: int = None

@app.post("/internal/generate-plans")
async def generate_plans(req: PlanRequest):
    from planner_service import generate_plans
    count = await generate_plans(
        req.class_id, req.section_id, req.school_id, req.academic_year,
        month=req.month, plan_year=req.plan_year
    )
    return {"plans_created": count}


# --- Coverage analysis ---

class CoverageRequest(BaseModel):
    coverage_log_id: str
    log_text: str
    section_id: str
    log_date: str

@app.post("/internal/analyze-coverage")
async def analyze_coverage(req: CoverageRequest):
    from coverage_analyzer import analyze_coverage
    results = await analyze_coverage(req.coverage_log_id, req.log_text, req.section_id, req.log_date)
    return {"results": results}


# --- Holiday import ---

@app.post("/internal/import-holidays")
async def import_holidays(file: UploadFile = File(...)):
    import openpyxl
    content = await file.read()
    valid_rows = []
    invalid_rows = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

        # Flexible column detection — support multiple naming conventions
        def find_col(candidates):
            for c in candidates:
                if c in headers:
                    return headers.index(c)
            return None

        date_idx = find_col(['date'])
        # event_name can come from 'description', 'event_name', 'event', 'name', 'holiday name'
        name_idx = find_col(['description', 'event_name', 'event', 'name', 'holiday name', 'holiday'])
        # type column: 'type', 'day type', 'category'
        type_idx = find_col(['type', 'day type', 'category'])

        if date_idx is None:
            raise HTTPException(status_code=400, detail="Missing required column: Date")
        if name_idx is None:
            raise HTTPException(status_code=400, detail="Missing required column: Description or event_name")

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_date = row[date_idx] if date_idx < len(row) else None
            event_name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
            row_type = str(row[type_idx]).strip().lower() if type_idx is not None and type_idx < len(row) and row[type_idx] else ''

            if not event_name or not str(event_name).strip():
                invalid_rows.append({"row": row_num, "reason": "Missing description/event name"})
                continue

            # Skip rows that are explicitly "Working Day" — those aren't holidays
            if row_type in ('working day', 'working', 'school day'):
                invalid_rows.append({"row": row_num, "reason": f"Skipped: type is '{row_type}' (not a holiday)"})
                continue

            # Parse date
            parsed_date = None
            if isinstance(raw_date, (datetime, date)):
                parsed_date = raw_date.strftime('%Y-%m-%d') if isinstance(raw_date, datetime) else str(raw_date)
            elif isinstance(raw_date, str):
                raw_str = raw_date.strip()
                for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%d %b %Y', '%d %B %Y', '%B %d %Y'):
                    try:
                        parsed_date = datetime.strptime(raw_str, fmt).strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        continue

            if not parsed_date:
                invalid_rows.append({"row": row_num, "reason": f"Invalid date: {raw_date}"})
                continue

            valid_rows.append({
                "date": parsed_date,
                "event_name": str(event_name).strip(),
                "type": row_type or "holiday"
            })

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    return {"valid_rows": valid_rows, "invalid_rows": invalid_rows}


# --- Student import ---

REQUIRED_STUDENT_COLUMNS = {'student name', 'father name', 'section', 'class', 'parent contact number'}

@app.post("/internal/import-students")
async def import_students(file: UploadFile = File(...)):
    import openpyxl
    content = await file.read()
    valid_rows = []
    invalid_rows = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

        missing = REQUIRED_STUDENT_COLUMNS - set(headers)
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(sorted(missing))}")

        col = {h: i for i, h in enumerate(headers)}

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            def get(key):
                idx = col.get(key)
                return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ''

            student_name = get('student name')
            father_name = get('father name')
            section = get('section')
            class_name = get('class')
            parent_contact = get('parent contact number')

            if not student_name:
                invalid_rows.append({"row": row_num, "reason": "Missing student name"})
                continue

            valid_rows.append({
                "student_name": student_name,
                "father_name": father_name,
                "section": section,
                "class": class_name,
                "parent_contact": parent_contact,
            })

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    return {"valid_rows": valid_rows, "invalid_rows": invalid_rows}


# --- Holiday list PDF export ---

class HolidayExportRequest(BaseModel):
    academic_year: str
    holidays: list

@app.post("/internal/export-holiday-pdf")
async def export_holiday_pdf(req: HolidayExportRequest):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.colors import Color
    from datetime import date as date_type
    import io

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    DAY_NAMES = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']

    def draw_watermark(canvas):
        canvas.saveState()
        canvas.setFont("Helvetica", 60)
        canvas.setFillColor(Color(0.85, 0.85, 0.85, alpha=0.3))
        canvas.translate(width / 2, height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "OAKIT.AI")
        canvas.restoreState()

    draw_watermark(c)
    y = height - 2 * cm

    c.setFont("Helvetica-Bold", 16)
    c.drawString(2 * cm, y, f"Holiday List — {req.academic_year}")
    y -= 0.6 * cm
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    from datetime import datetime
    c.drawString(2 * cm, y, f"Generated on {datetime.now().strftime('%d %B %Y')}   ·   {len(req.holidays)} holidays")
    c.setFillColorRGB(0, 0, 0)
    y -= 0.4 * cm
    c.line(2 * cm, y, width - 2 * cm, y)
    y -= 0.7 * cm

    # Table header
    c.setFont("Helvetica-Bold", 9)
    c.setFillColorRGB(0.3, 0.3, 0.3)
    c.drawString(2 * cm, y, "#")
    c.drawString(2.8 * cm, y, "Holiday")
    c.drawString(12 * cm, y, "Day")
    c.drawString(14.5 * cm, y, "Date")
    c.setFillColorRGB(0, 0, 0)
    y -= 0.3 * cm
    c.line(2 * cm, y, width - 2 * cm, y)
    y -= 0.5 * cm

    for i, h in enumerate(req.holidays, 1):
        if y < 3 * cm:
            c.showPage()
            draw_watermark(c)
            y = height - 2 * cm

        try:
            d = date_type.fromisoformat(str(h.get('date', ''))[:10])
            day_name = DAY_NAMES[d.weekday()]
            date_str = d.strftime('%d %b %Y')
        except Exception:
            day_name = ''
            date_str = str(h.get('date', ''))[:10]

        # Alternating row background
        if i % 2 == 0:
            c.setFillColorRGB(0.97, 0.97, 0.97)
            c.rect(1.8 * cm, y - 0.15 * cm, width - 3.6 * cm, 0.55 * cm, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)

        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0.5, 0.5, 0.5)
        c.drawString(2 * cm, y, str(i))
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(2.8 * cm, y, h.get('event_name', '')[:55])
        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0.2, 0.5, 0.3)
        c.drawString(12 * cm, y, day_name)
        c.setFillColorRGB(0.3, 0.3, 0.3)
        c.drawString(14.5 * cm, y, date_str)
        c.setFillColorRGB(0, 0, 0)
        y -= 0.55 * cm

    c.save()
    buf.seek(0)
    from fastapi.responses import Response as FastResponse
    return FastResponse(content=buf.read(), media_type="application/pdf")


# --- Monthly plan PDF export ---

class MonthlyExportRequest(BaseModel):
    section_id: str
    section_label: str
    class_name: str
    month_name: str
    plans: list = []
    days: list = []

@app.post("/internal/export-monthly-pdf")
async def export_monthly_pdf(req: MonthlyExportRequest):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.colors import Color
    import io

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    def draw_watermark(canvas):
        canvas.saveState()
        canvas.setFont("Helvetica", 60)
        canvas.setFillColor(Color(0.85, 0.85, 0.85, alpha=0.3))
        canvas.translate(width / 2, height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "OAKIT.AI")
        canvas.restoreState()

    draw_watermark(c)
    y = height - 2 * cm

    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(2 * cm, y, f"Monthly Plan — {req.month_name}")
    y -= 0.7 * cm
    c.setFont("Helvetica", 11)
    c.drawString(2 * cm, y, f"Class: {req.class_name}   |   Section: {req.section_label}")
    y -= 0.5 * cm
    c.line(2 * cm, y, width - 2 * cm, y)
    y -= 0.7 * cm

    for plan in (req.plans or []):
        if y < 4 * cm:
            c.showPage()
            draw_watermark(c)
            y = height - 2 * cm

        day_type = plan.get('type', 'working')
        plan_date = plan.get('date') or plan.get('plan_date', '')
        try:
            from datetime import date as date_type
            d = date_type.fromisoformat(str(plan_date)[:10])
            date_str = d.strftime('%a, %d %b')
        except Exception:
            date_str = str(plan_date)[:10]

        if day_type == 'weekend':
            c.setFont("Helvetica", 9)
            c.setFillColorRGB(0.7, 0.7, 0.7)
            c.drawString(2 * cm, y, f"{date_str}   —   Weekend")
            c.setFillColorRGB(0, 0, 0)
            y -= 0.45 * cm

        elif day_type == 'holiday':
            c.setFont("Helvetica-Bold", 9)
            c.setFillColorRGB(0.8, 0.1, 0.1)
            holiday_label = plan.get('label', 'Holiday')
            c.drawString(2 * cm, y, f"{date_str}   \U0001f389  {holiday_label}")
            c.setFillColorRGB(0, 0, 0)
            y -= 0.45 * cm

        else:
            c.setFont("Helvetica-Bold", 10)
            c.setFillColorRGB(0.1, 0.3, 0.2)
            c.drawString(2 * cm, y, date_str)
            c.setFillColorRGB(0, 0, 0)
            y -= 0.45 * cm

            chunks = plan.get('chunks', [])
            if not chunks:
                c.setFont("Helvetica-Oblique", 8)
                c.setFillColorRGB(0.6, 0.6, 0.6)
                c.drawString(2.5 * cm, y, "No plan assigned")
                c.setFillColorRGB(0, 0, 0)
                y -= 0.4 * cm
            else:
                for chunk in chunks:
                    if y < 3 * cm:
                        c.showPage()
                        draw_watermark(c)
                        y = height - 2 * cm
                    c.setFont("Helvetica-Bold", 9)
                    topic = chunk.get('topic_label', 'Topic')
                    c.drawString(2.5 * cm, y, f"• {topic}")
                    y -= 0.4 * cm

                    content = chunk.get('content', '').strip()
                    if content:
                        for line in content.split('\n'):
                            line = line.strip()
                            if not line:
                                continue
                            if y < 3 * cm:
                                c.showPage()
                                draw_watermark(c)
                                y = height - 2 * cm
                            if ':' in line:
                                parts = line.split(':', 1)
                                subject = parts[0].strip()
                                activity = parts[1].strip()
                                c.setFont("Helvetica-Bold", 8)
                                c.setFillColorRGB(0.3, 0.3, 0.3)
                                c.drawString(3 * cm, y, f"{subject}:")
                                c.setFont("Helvetica", 8)
                                c.setFillColorRGB(0, 0, 0)
                                max_chars = 80
                                if len(activity) > max_chars:
                                    c.drawString(5.5 * cm, y, activity[:max_chars])
                                    y -= 0.35 * cm
                                    c.drawString(5.5 * cm, y, activity[max_chars:max_chars*2])
                                else:
                                    c.drawString(5.5 * cm, y, activity)
                            else:
                                c.setFont("Helvetica", 8)
                                c.drawString(3 * cm, y, line[:100])
                            y -= 0.35 * cm

                    activity_ids = chunk.get('activity_ids', [])
                    if activity_ids:
                        if y < 3 * cm:
                            c.showPage()
                            draw_watermark(c)
                            y = height - 2 * cm
                        c.setFont("Helvetica-Oblique", 7.5)
                        c.setFillColorRGB(0.2, 0.5, 0.3)
                        c.drawString(3 * cm, y, f"\U0001f4ce {', '.join(activity_ids)}")
                        c.setFillColorRGB(0, 0, 0)
                        y -= 0.35 * cm

                    y -= 0.15 * cm

        y -= 0.1 * cm

    c.save()
    buf.seek(0)
    from fastapi.responses import Response as FastResponse
    return FastResponse(content=buf.read(), media_type="application/pdf")


# --- PDF export ---

class ExportPdfRequest(BaseModel):
    teacher_id: str
    section_id: str
    section_label: str = ""
    teacher_name: str = "Teacher"
    date: str
    days: int = 1
    settling_text: str | None = None

@app.post("/internal/export-pdf")
async def export_pdf(req: ExportPdfRequest):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.colors import Color
    from db import get_pool
    from uuid import UUID
    from datetime import date as date_type, timedelta

    pool = await get_pool()
    start_date = date_type.fromisoformat(req.date)

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    def draw_watermark(canvas):
        canvas.saveState()
        canvas.setFont("Helvetica", 60)
        canvas.setFillColor(Color(0.85, 0.85, 0.85, alpha=0.3))
        canvas.translate(width / 2, height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "OAKIT.AI")
        canvas.restoreState()

    # If settling_text provided — render the AI-generated plan as PDF
    if req.settling_text:
        draw_watermark(c)
        y = height - 2 * cm
        c.setFont("Helvetica-Bold", 14)
        c.drawString(2 * cm, y, f"Oakie's Daily Planner — {start_date.strftime('%d %B %Y')}")
        y -= 0.7 * cm
        c.setFont("Helvetica", 11)
        c.drawString(2 * cm, y, f"Teacher: {req.teacher_name}   |   Section: {req.section_label}")
        y -= 0.5 * cm
        c.line(2 * cm, y, width - 2 * cm, y)
        y -= 0.7 * cm

        # Render each line of the settling text
        import re
        for line in req.settling_text.split('\n'):
            line = line.strip()
            if not line:
                y -= 0.3 * cm
                continue
            if y < 3 * cm:
                c.showPage()
                draw_watermark(c)
                y = height - 2 * cm
            # Strip emoji for PDF (reportlab doesn't support all emoji)
            clean = re.sub(r'[^\x00-\x7F]+', '', line).strip()
            if not clean:
                continue
            if line.startswith('📌') or line.startswith('What to do') or line.startswith('Tip'):
                c.setFont("Helvetica-Bold", 10)
            else:
                c.setFont("Helvetica", 9)
            c.drawString(2 * cm, y, clean[:100])
            y -= 0.45 * cm

        c.save()
        buf.seek(0)
        return Response(content=buf.read(), media_type="application/pdf")

    # Regular curriculum day export
    plan = await pool.fetchrow(
        "SELECT plan_date, chunk_ids, chunk_label_overrides FROM day_plans WHERE section_id = $1 AND plan_date = $2",
        UUID(req.section_id), start_date
    )
    if not plan or not plan['chunk_ids']:
        raise HTTPException(status_code=404, detail="No day plan found for the requested date")

    overrides = dict(plan['chunk_label_overrides']) if plan['chunk_label_overrides'] else {}

    chunks = await pool.fetch(
        "SELECT id, topic_label, content, activity_ids FROM curriculum_chunks WHERE id = ANY($1::uuid[]) ORDER BY chunk_index",
        plan['chunk_ids']
    )
    # Apply label overrides
    chunks_with_overrides = []
    for chunk in chunks:
        label = overrides.get(str(chunk['id']), chunk['topic_label'])
        chunks_with_overrides.append({**dict(chunk), 'topic_label': label})

    draw_watermark(c)
    y = height - 2 * cm

    c.setFont("Helvetica-Bold", 14)
    c.drawString(2 * cm, y, f"Oakie's Daily Planner — {start_date.strftime('%d %B %Y')}")
    y -= 0.7 * cm
    c.setFont("Helvetica", 11)
    c.drawString(2 * cm, y, f"Teacher: {req.teacher_name}   |   Section: {req.section_label}")
    y -= 0.5 * cm
    c.line(2 * cm, y, width - 2 * cm, y)
    y -= 0.7 * cm

    for chunk in chunks_with_overrides:
        if y < 3 * cm:
            c.showPage()
            draw_watermark(c)
            y = height - 2 * cm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(2 * cm, y, f"• {chunk['topic_label'] or 'Topic'}")
        y -= 0.5 * cm

        content = (chunk['content'] or '').strip()
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue
            if y < 3 * cm:
                c.showPage()
                draw_watermark(c)
                y = height - 2 * cm
            c.setFont("Helvetica", 9)
            c.drawString(2.5 * cm, y, line[:100])
            y -= 0.4 * cm

        if chunk.get('activity_ids'):
            c.setFont("Helvetica-Oblique", 9)
            c.setFillColorRGB(0.2, 0.5, 0.3)
            c.drawString(2.5 * cm, y, f"Materials: {', '.join(chunk['activity_ids'])}")
            c.setFillColorRGB(0, 0, 0)
            y -= 0.4 * cm

        y -= 0.3 * cm

    c.save()
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf")


# --- Progress Report PDF export ---

class ProgressReportPdfRequest(BaseModel):
    student_name: str
    age: str = ""
    class_name: str
    section_label: str
    teacher_name: str = ""
    father_name: str = ""
    mother_name: str = ""
    school_name: str
    from_date: str
    to_date: str
    attendance_pct: int = 0
    attendance_present: int = 0
    attendance_total: int = 0
    curriculum_covered: int = 0
    milestones_achieved: int = 0
    milestones_total: int = 0
    homework_completed: int = 0
    homework_total: int = 0
    ai_report: str

@app.post("/internal/export-progress-report-pdf")
async def export_progress_report_pdf(req: ProgressReportPdfRequest):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.colors import HexColor, Color, white, black
    from reportlab.platypus import Paragraph, Frame
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    import io, re, textwrap

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    BRAND = HexColor('#1A3C2E')
    BRAND_LIGHT = HexColor('#2E7D5E')
    NEUTRAL = HexColor('#9E9690')
    LIGHT_BG = HexColor('#F5F4F2')
    BORDER = HexColor('#ECEAE7')
    GREEN_BG = HexColor('#F0FDF4')
    GREEN_TEXT = HexColor('#15803D')
    AMBER_BG = HexColor('#FFFBEB')
    AMBER_TEXT = HexColor('#B45309')
    BLUE_BG = HexColor('#EFF6FF')
    BLUE_TEXT = HexColor('#1D4ED8')

    margin = 1.8 * cm
    col_w = width - 2 * margin

    def new_page():
        c.showPage()
        # Subtle header on continuation pages
        c.setFillColor(BRAND)
        c.rect(0, height - 0.8*cm, width, 0.8*cm, fill=1, stroke=0)
        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(margin, height - 0.55*cm, f"{req.student_name} — Progress Report · {req.school_name}")
        return height - 1.5*cm

    # ── Header ──────────────────────────────────────────────────
    c.setFillColor(BRAND)
    c.rect(0, height - 4.5*cm, width, 4.5*cm, fill=1, stroke=0)
    # Subtle gradient overlay
    c.setFillColor(BRAND_LIGHT)
    c.setFillAlpha(0.3)
    c.rect(width*0.6, height - 4.5*cm, width*0.4, 4.5*cm, fill=1, stroke=0)
    c.setFillAlpha(1.0)

    c.setFillColor(white)
    c.setFont("Helvetica", 8)
    c.setFillAlpha(0.65)
    c.drawString(margin, height - 1.2*cm, req.school_name.upper())
    c.setFillAlpha(1.0)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(margin, height - 2.1*cm, f"Progress Report — {req.student_name}")
    c.setFont("Helvetica", 10)
    c.setFillAlpha(0.75)
    c.drawString(margin, height - 2.8*cm, f"{req.from_date}  to  {req.to_date}")
    c.setFillAlpha(1.0)
    c.setFont("Helvetica", 9)
    c.setFillAlpha(0.6)
    c.drawString(margin, height - 3.5*cm, f"Generated on {datetime.now().strftime('%d %B %Y')}")
    c.setFillAlpha(1.0)

    y = height - 5.2*cm

    # ── Profile strip ────────────────────────────────────────────
    c.setFillColor(LIGHT_BG)
    c.roundRect(margin, y - 1.4*cm, col_w, 1.6*cm, 6, fill=1, stroke=0)
    fields = [
        ("Student", f"{req.student_name}{' (' + req.age + ')' if req.age else ''}"),
        ("Class", f"{req.class_name} · {req.section_label}"),
        ("Teacher", req.teacher_name or "—"),
        ("Father", req.father_name or "—"),
    ]
    fw = col_w / 4
    for i, (lbl, val) in enumerate(fields):
        x = margin + i * fw + 0.3*cm
        c.setFont("Helvetica", 7)
        c.setFillColor(NEUTRAL)
        c.drawString(x, y - 0.5*cm, lbl.upper())
        c.setFont("Helvetica-Bold", 9)
        c.setFillColor(black)
        # Truncate long values
        v = val[:22] if len(val) > 22 else val
        c.drawString(x, y - 1.0*cm, v)
    y -= 2.0*cm

    # ── Stat cards ───────────────────────────────────────────────
    stats = [
        (f"{req.attendance_pct}%", f"Attendance\n{req.attendance_present}/{req.attendance_total} days",
         GREEN_BG if req.attendance_pct >= 90 else AMBER_BG, GREEN_TEXT if req.attendance_pct >= 90 else AMBER_TEXT),
        (str(req.curriculum_covered), "Topics\nCovered", BLUE_BG, BLUE_TEXT),
        (f"{req.milestones_achieved}/{req.milestones_total}", "Milestones\nAchieved", AMBER_BG, AMBER_TEXT),
        (f"{req.homework_completed}/{req.homework_total}", "Homework\nCompleted", LIGHT_BG, NEUTRAL),
    ]
    sw = (col_w - 0.3*cm * 3) / 4
    for i, (val, lbl, bg, fg) in enumerate(stats):
        sx = margin + i * (sw + 0.3*cm)
        c.setFillColor(bg)
        c.roundRect(sx, y - 1.5*cm, sw, 1.7*cm, 5, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 16)
        c.setFillColor(fg)
        c.drawCentredString(sx + sw/2, y - 0.8*cm, val)
        c.setFont("Helvetica", 7.5)
        c.setFillColor(NEUTRAL)
        for j, line in enumerate(lbl.split('\n')):
            c.drawCentredString(sx + sw/2, y - 1.15*cm - j*0.28*cm, line)
    y -= 2.2*cm

    # ── Report sections ──────────────────────────────────────────
    clean = re.sub(r'\*\*([^*]+)\*\*', r'\1', req.ai_report)
    clean = re.sub(r'\*([^*]+)\*', r'\1', clean)
    sections = [s for s in re.split(r'\n##\s+', clean) if s.strip()]

    for section in sections:
        lines = section.split('\n')
        heading = lines[0].strip()
        body_lines = [l.strip() for l in lines[1:] if l.strip()]

        # Estimate height needed
        est_h = 0.7*cm + len(body_lines) * 0.45*cm + 0.4*cm
        if y - est_h < 2.5*cm:
            y = new_page()

        # Section box
        box_h = 0.65*cm + len(body_lines) * 0.48*cm + 0.3*cm
        c.setFillColor(LIGHT_BG)
        c.roundRect(margin, y - box_h, col_w, box_h, 5, fill=1, stroke=0)
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.5)
        c.roundRect(margin, y - box_h, col_w, box_h, 5, fill=0, stroke=1)

        # Heading
        c.setFillColor(black)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(margin + 0.4*cm, y - 0.5*cm, heading)
        c.setStrokeColor(BORDER)
        c.line(margin, y - 0.7*cm, margin + col_w, y - 0.7*cm)

        # Body lines
        by = y - 1.05*cm
        for line in body_lines:
            if by < 2.5*cm:
                y = new_page()
                by = y - 0.3*cm
            # Check for "Label: content" pattern
            sub = re.match(r'^([A-Za-z\s&/]+?):\s+(.+)', line)
            if sub and len(sub.group(1)) < 35:
                c.setFont("Helvetica-Bold", 8.5)
                c.setFillColor(NEUTRAL)
                c.drawString(margin + 0.4*cm, by, sub.group(1) + ":")
                c.setFont("Helvetica", 8.5)
                c.setFillColor(black)
                # Wrap long text
                wrapped = textwrap.fill(sub.group(2), width=72)
                for wl in wrapped.split('\n'):
                    c.drawString(margin + 4.5*cm, by, wl[:90])
                    by -= 0.42*cm
            else:
                c.setFont("Helvetica", 8.5)
                c.setFillColor(black)
                wrapped = textwrap.fill(line, width=95)
                for wl in wrapped.split('\n'):
                    c.drawString(margin + 0.4*cm, by, wl[:100])
                    by -= 0.42*cm
            by -= 0.06*cm

        y = by - 0.4*cm

    # ── Footer ───────────────────────────────────────────────────
    if y < 2.5*cm:
        c.showPage()
        y = height - 2*cm
    c.setStrokeColor(BORDER)
    c.line(margin, 1.8*cm, margin + col_w, 1.8*cm)
    c.setFont("Helvetica", 7.5)
    c.setFillColor(NEUTRAL)
    c.drawCentredString(width/2, 1.3*cm, f"Generated by Oakit.ai  ·  {req.school_name}  ·  {datetime.now().strftime('%d %B %Y')}")

    c.save()
    buf.seek(0)
    from fastapi.responses import Response as FastResponse
    fname = f"Progress_Report_{req.student_name.replace(' ', '_')}.pdf"
    return FastResponse(
        content=buf.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# --- Greeting ---

@app.get("/internal/greeting")
async def greeting(teacher_name: str = "Teacher", teacher_id: str = ""):
    now = datetime.now()
    hour = now.hour

    if 5 <= hour < 12:
        salutation = "Good morning"
    elif 12 <= hour < 17:
        salutation = "Good afternoon"
    else:
        salutation = "Good evening"

    greeting_text = f"{salutation}, {teacher_name}!"

    # Early arrival (before 7am)
    if hour < 7:
        greeting_text += " You're here early — that's dedication!"

    # Rotate thought by date hash
    date_str = now.strftime('%Y-%m-%d') + teacher_id
    idx = int(hashlib.md5(date_str.encode()).hexdigest(), 16) % len(THOUGHTS_FOR_DAY)
    thought = THOUGHTS_FOR_DAY[idx]

    return {
        "greeting": greeting_text,
        "thought_for_day": thought,
    }


# --- Birthday wish generator ---

class BirthdayWishRequest(BaseModel):
    students: list  # [{name, age, class_name, section_label}]
    school_name: str = ""

@app.post("/internal/birthday-wish")
async def birthday_wish(req: BirthdayWishRequest):
    """Generate warm, age-appropriate birthday wishes for students turning X today."""
    from llm_client import call_llm

    if not req.students:
        return {"wishes": {}}

    # Build a prompt describing each child
    lines = []
    for s in req.students:
        age = s.get("age")
        name = s.get("name", "Student")
        cls = s.get("class_name", "")
        age_str = f", turning {age} today" if age else ""
        lines.append(f"- {name} ({cls}{age_str})")

    students_text = "\n".join(lines)
    today_str = datetime.now().strftime("%d %B %Y")

    prompt = f"""Today is {today_str}. The following students are celebrating their birthday today at school:

{students_text}

Write a single warm, joyful birthday announcement message (3-4 sentences) that:
- Celebrates all these children together
- Is age-appropriate and cheerful for a school setting
- Mentions the school community wishing them well
- Ends with an encouraging note about their learning journey
- Uses simple, warm language suitable for parents to read

Return ONLY the message text, no extra formatting."""

    try:
        result = await call_llm(prompt, max_tokens=200)
        message = result.strip() if isinstance(result, str) else result.get("text", "").strip()
    except Exception:
        # Fallback message
        names = ", ".join(s.get("name", "Student") for s in req.students)
        message = f"🎂 Wishing a very Happy Birthday to {names}! May this special day be filled with joy, laughter, and wonderful memories. The entire school family celebrates with you today. Keep shining bright in your learning journey! 🌟"

    return {"message": message, "student_count": len(req.students)}


class HomeworkFormatRequest(BaseModel):
    raw_text: str
    school_id: str = ""
    section_id: str = ""

class GenerateReportRequest(BaseModel):
    prompt: str
    student_name: str = ""

@app.post("/internal/generate-report")
async def generate_report(req: GenerateReportRequest):
    """Generate a student progress report — bypasses query pipeline filters.
    Called directly by the admin reports endpoint."""
    from query_pipeline import _call_llm

    system = (
        f"You are an expert school report writer generating a formal progress report for {req.student_name or 'a student'}.\n"
        "Write in warm, professional, parent-friendly language.\n"
        "Use the section headings provided (## heading). Write flowing sentences — NO bullet points, NO bold markdown, NO asterisks.\n"
        "Be specific, positive, and encouraging. Keep each section to 2-4 sentences.\n"
        "Always use the student's name — never 'the child' or 'your child'."
    )

    try:
        response, provider = await _call_llm(req.prompt, system)
        if not response:
            response = f"Progress report for {req.student_name} could not be generated at this time. Please try again."
        # Strip any markdown bold that slipped through
        import re
        response = re.sub(r'\*\*([^*]+)\*\*', r'\1', response)
        response = re.sub(r'\*([^*]+)\*', r'\1', response)
    except Exception as e:
        response = f"Report generation failed: {str(e)}"

    return {"response": response}


@app.post("/internal/format-homework")
async def format_homework(req: HomeworkFormatRequest):
    """Format raw teacher homework text into a clean, parent-friendly message.
    Also checks relevance against today's covered topics and adds a note if mismatched.
    """
    from query_pipeline import _call_llm
    from db import get_pool

    # Fetch today's covered topics for relevance check
    covered_topics: list[str] = []
    try:
        pool = await get_pool()
        rows = await pool.fetch(
            """SELECT cc.topic_label, cc.content
               FROM daily_completions dc
               JOIN curriculum_chunks cc ON cc.id = ANY(dc.covered_chunk_ids)
               WHERE dc.section_id = $1
               ORDER BY dc.completion_date DESC, cc.chunk_index
               LIMIT 20""",
            req.section_id,
        )
        covered_topics = [r["topic_label"] for r in rows if r["topic_label"]]
    except Exception:
        pass  # non-critical — proceed without relevance check

    topics_context = ""
    if covered_topics:
        topics_context = f"\nTopics covered today: {', '.join(covered_topics[:8])}"

    system = (
        "You are formatting a homework message from a teacher to parents of preschool/primary school children.\n"
        "Output plain text only — no markdown, no bold, no bullet symbols.\n"
        "Keep it warm, clear, and concise. Under 150 words.\n"
        "Start with 'Homework for today:' then list the tasks clearly numbered.\n"
        "If the homework is relevant to today's topics, add a short line connecting them (e.g. 'This reinforces what we learned about...').\n"
        "End with one short encouraging line for parents."
    )
    prompt = f"""Teacher's raw homework note:
\"{req.raw_text}\"{topics_context}

Rewrite this as a clear, friendly homework message for parents.
Format each task as a numbered item.
Keep the original tasks — do not add or remove any.
If the homework relates to the topics covered today, add one sentence connecting them.
Under 150 words."""

    try:
        formatted, _ = await _call_llm(prompt, system)
        if not formatted:
            formatted = f"Homework for today:\n{req.raw_text}"
    except Exception:
        formatted = f"Homework for today:\n{req.raw_text}"

    return {"formatted_text": formatted}


# --- Child Journey beautifier ---

class ChildJourneyRequest(BaseModel):
    raw_text: str
    student_name: str = ""
    class_level: str = ""
    entry_type: str = "daily"  # daily, weekly, highlight
    entry_date: str = ""

@app.post("/internal/beautify-child-journey")
async def beautify_child_journey(req: ChildJourneyRequest):
    """
    Transform a teacher's short raw notes about a child into a warm,
    parent-friendly narrative. Keeps it personal, positive, and concise.
    """
    from query_pipeline import _call_llm

    entry_label = {
        "daily": "today",
        "weekly": "this week",
        "highlight": "a special moment",
    }.get(req.entry_type, "today")

    system = (
        f"You are writing a warm, personal update about a child for their parents.\n"
        f"The child's name is {req.student_name or 'the child'}.\n"
        f"Write in second person to the parents (e.g. 'Aarav showed...' or 'Your child...').\n"
        f"Keep it warm, specific, encouraging, and under 100 words.\n"
        f"Plain text only — no markdown, no bullet points.\n"
        f"Focus on what the child did, showed, or achieved — not generic praise."
    )

    prompt = (
        f"Teacher's notes about {req.student_name or 'the child'} for {entry_label}"
        f"{' (' + req.entry_date + ')' if req.entry_date else ''}:\n\n"
        f"\"{req.raw_text}\"\n\n"
        f"Rewrite this as a warm, personal update for the parents. "
        f"Keep it specific to what the teacher observed. Under 100 words."
    )

    try:
        beautified, _ = await _call_llm(prompt, system)
        if not beautified:
            beautified = req.raw_text
    except Exception:
        beautified = req.raw_text

    return {"beautified_text": beautified}


# --- Activity suggestions ---

class SuggestActivityRequest(BaseModel):
    subject: str
    class_level: str
    topic: str = ""
    school_id: str = ""
    query_date: str = ""

@app.post("/internal/suggest-activity")
async def suggest_activity(req: SuggestActivityRequest):
    """Generate differentiated activity suggestions for a subject/topic."""
    from query_pipeline import _call_llm

    system = (
        "You are an expert early childhood educator helping a teacher plan activities.\n"
        "Generate exactly 3 activity ideas: Simple, Standard, and Extended difficulty.\n"
        "Each activity must be age-appropriate for the class level specified.\n"
        "Return JSON only — no markdown, no explanation outside the JSON."
    )
    prompt = f"""Subject: {req.subject}
Class Level: {req.class_level}
Topic: {req.topic or req.subject}

Generate 3 differentiated activity ideas as JSON:
{{
  "activities": [
    {{
      "title": "short activity name",
      "description": "2-3 sentence description of what to do",
      "difficulty": "Simple",
      "support_level": "additional"
    }},
    {{
      "title": "short activity name",
      "description": "2-3 sentence description",
      "difficulty": "Standard",
      "support_level": "standard"
    }},
    {{
      "title": "short activity name",
      "description": "2-3 sentence description",
      "difficulty": "Extended",
      "support_level": "advanced"
    }}
  ]
}}"""

    try:
        import json, asyncio
        result, _ = await _call_llm(prompt, system)
        # Extract JSON from response
        start = result.find('{')
        end = result.rfind('}') + 1
        if start >= 0 and end > start:
            return json.loads(result[start:end])
    except Exception:
        pass

    # Fallback
    return {
        "activities": [
            {"title": f"{req.subject} — Simple", "description": f"A simple {req.subject.lower()} activity for {req.class_level}. Use visual aids and repetition.", "difficulty": "Simple", "support_level": "additional"},
            {"title": f"{req.subject} — Standard", "description": f"A standard {req.subject.lower()} activity. Guide children through the topic with hands-on materials.", "difficulty": "Standard", "support_level": "standard"},
            {"title": f"{req.subject} — Extended", "description": f"An extended activity for advanced {req.class_level} learners. Encourage creative expression and explanation.", "difficulty": "Extended", "support_level": "advanced"},
        ]
    }


# --- Worksheet generator ---

class WorksheetRequest(BaseModel):
    subject: str
    topic: str = ""
    class_level: str

@app.post("/internal/generate-worksheet")
async def generate_worksheet(req: WorksheetRequest):
    """Generate a printable worksheet for a subject/topic."""
    from query_pipeline import _call_llm
    import json

    system = (
        "You are creating a simple printable worksheet for preschool/primary children.\n"
        "The worksheet must be age-appropriate for the class level.\n"
        "Return JSON only — no markdown outside the JSON."
    )
    prompt = f"""Create a worksheet for:
Subject: {req.subject}
Topic: {req.topic or req.subject}
Class Level: {req.class_level}

Return JSON:
{{
  "title": "worksheet title",
  "topic": "{req.topic or req.subject}",
  "class_level": "{req.class_level}",
  "sections": [
    {{
      "type": "fill_in_blank",
      "title": "Fill in the Blanks",
      "items": [
        {{"question": "The ___ is red.", "answer": "apple"}},
        {{"question": "We have ___ fingers.", "answer": "ten"}}
      ]
    }},
    {{
      "type": "matching",
      "title": "Match the Following",
      "items": [
        {{"left": "Sun", "right": "Shines in the day"}},
        {{"left": "Moon", "right": "Shines at night"}}
      ]
    }},
    {{
      "type": "drawing",
      "title": "Draw and Colour",
      "prompt": "Draw a picture related to the topic and colour it."
    }}
  ]
}}"""

    try:
        result, _ = await _call_llm(prompt, system)
        start = result.find('{')
        end = result.rfind('}') + 1
        if start >= 0 and end > start:
            return json.loads(result[start:end])
    except Exception:
        pass

    return {
        "title": f"{req.subject} Worksheet",
        "topic": req.topic or req.subject,
        "class_level": req.class_level,
        "no_curriculum_content": True,
        "sections": [
            {"type": "fill_in_blank", "title": "Fill in the Blanks", "items": [{"question": "The sky is ___.", "answer": "blue"}, {"question": "A cat says ___.", "answer": "meow"}]},
            {"type": "matching", "title": "Match the Following", "items": [{"left": "Sun", "right": "Day"}, {"left": "Moon", "right": "Night"}]},
            {"type": "drawing", "title": "Draw and Colour", "prompt": "Draw your favourite animal and colour it."},
        ],
    }


# ─── Student Portal AI Endpoints ─────────────────────────────────────────────

class StudentQueryRequest(BaseModel):
    student_id: str
    school_id: str
    section_id: str
    text: str
    covered_chunk_ids: list
    query_date: str

@app.post("/internal/student-query")
async def student_query(req: StudentQueryRequest):
    """AI doubts scoped strictly to covered topics."""
    from query_pipeline import _call_llm, _build_chunk_context
    from db import get_pool
    from uuid import UUID

    pool = await get_pool()

    if not req.covered_chunk_ids:
        return {"response": "Your class hasn't covered any topics yet. Check back after your teacher logs some completed lessons! 🌱"}

    # Fetch covered chunks
    chunks = await pool.fetch(
        "SELECT id, topic_label, content FROM curriculum_chunks WHERE id = ANY($1::uuid[]) ORDER BY chunk_index",
        [UUID(c) for c in req.covered_chunk_ids[:50]],  # cap at 50
    )
    if not chunks:
        return {"response": "I couldn't find your covered topics. Please try again later."}

    # Build topic list for relevance check
    topic_labels = [c["topic_label"] or "Topic" for c in chunks]
    topics_text = "\n".join(f"- {t}" for t in topic_labels[:20])

    # Check relevance + generate answer in one LLM call
    sys_p = (
        "You are Oakie, a friendly learning assistant for school children.\n"
        "You ONLY answer questions about topics the student's class has already covered.\n"
        "If the question is not about any covered topic, politely decline and suggest a covered topic.\n"
        "Give clear, simple, age-appropriate explanations. Use examples children can relate to.\n"
        "Plain text only — no markdown bold, no tables. Use emojis to make it friendly.\n"
        "Keep responses under 200 words."
    )
    llm_p = f"""Student's question: "{req.text}"

TOPICS COVERED BY THIS CLASS:
{topics_text}

CURRICULUM CONTENT (for reference):
{_build_chunk_context(list(chunks[:10]))}

If the question is about a covered topic:
- Give a clear, friendly explanation
- Use a simple example
- End with an encouraging line

If the question is NOT about any covered topic:
- Say: "I can only help with topics your class has covered."
- Suggest 2-3 relevant covered topics they could ask about instead

Answer:"""

    response_text, _ = await _call_llm(llm_p, sys_p)
    if not response_text:
        response_text = "Oakie is unavailable right now. Please try again shortly."

    return {"response": response_text}


class GenerateQuizRequest(BaseModel):
    quiz_id: str
    chunk_ids: list
    chunks: list  # [{ id, topic_label, content }]
    question_types: list  # ['fill_blank', '1_mark', '2_mark', 'descriptive']
    subject: str
    class_name: str

@app.post("/internal/generate-quiz")
async def generate_quiz(req: GenerateQuizRequest):
    """Generate quiz questions from confirmed topic list."""
    from query_pipeline import _call_llm
    import json, random

    if not req.chunks:
        return {"questions": []}

    # Build content context
    content_lines = []
    for chunk in req.chunks[:15]:
        label = chunk.get("topic_label", "Topic")
        content = (chunk.get("content") or "")[:300]
        content_lines.append(f"Topic: {label}\n{content}")
    content_text = "\n\n---\n\n".join(content_lines)

    q_types_str = ", ".join(req.question_types)
    n_questions = max(5, len(req.chunks) * 2)

    sys_p = (
        f"You are generating quiz questions for {req.class_name} students.\n"
        f"Questions must be specific to the curriculum content provided.\n"
        f"Make questions varied and random — avoid repeating the same phrasing.\n"
        f"Return ONLY valid JSON — no markdown, no explanation."
    )
    llm_p = f"""Generate {n_questions} quiz questions for subject: {req.subject}
Question types to include: {q_types_str}

CURRICULUM CONTENT:
{content_text}

Return JSON:
{{
  "questions": [
    {{
      "chunk_id": "<id from content above or null>",
      "subject": "{req.subject}",
      "question": "question text",
      "q_type": "fill_blank|1_mark|2_mark|descriptive",
      "marks": 1,
      "answer_key": "correct answer",
      "explanation": "brief explanation"
    }}
  ]
}}

Rules:
- fill_blank: sentence with ___ to fill in (1 mark)
- 1_mark: short answer question (1 mark)
- 2_mark: question requiring 2-3 sentences (2 marks)
- descriptive: open-ended question (2 marks)
- Make questions specific to the actual content, not generic
- Vary the difficulty and phrasing"""

    try:
        result, _ = await _call_llm(llm_p, sys_p)
        start = result.find('{')
        end = result.rfind('}') + 1
        if start >= 0 and end > start:
            data = json.loads(result[start:end])
            questions = data.get("questions", [])
            # Shuffle for randomness
            random.shuffle(questions)
            return {"questions": questions}
    except Exception as e:
        print(f"[generate-quiz] error: {e}")

    # Fallback: basic questions from topic labels
    questions = []
    for chunk in req.chunks:
        label = chunk.get("topic_label", "Topic")
        questions.append({
            "chunk_id": chunk.get("id"),
            "subject": req.subject,
            "question": f"What did you learn about {label}?",
            "q_type": "1_mark",
            "marks": 1,
            "answer_key": label,
            "explanation": f"This topic was covered in class.",
        })
    return {"questions": questions}


class EvaluateQuizRequest(BaseModel):
    questions: list  # [{ id, question, q_type, marks, answer_key, explanation }]
    student_answers: list  # [{ question_id, answer }]
    class_name: str

@app.post("/internal/evaluate-quiz")
async def evaluate_quiz(req: EvaluateQuizRequest):
    """Evaluate student answers and assign marks."""
    from query_pipeline import _call_llm
    import json

    q_map = {q["id"]: q for q in req.questions}
    answer_map = {a["question_id"]: a.get("answer", "") for a in req.student_answers}

    evaluations = []

    # Separate objective (fill_blank, 1_mark) from subjective (2_mark, descriptive)
    objective_ids = [q["id"] for q in req.questions if q["q_type"] in ("fill_blank", "1_mark")]
    subjective_ids = [q["id"] for q in req.questions if q["q_type"] in ("2_mark", "descriptive")]

    # Evaluate objective questions with fuzzy match
    for qid in objective_ids:
        q = q_map.get(qid)
        if not q:
            continue
        student_ans = (answer_map.get(qid) or "").strip().lower()
        correct_ans = (q.get("answer_key") or "").strip().lower()
        # Fuzzy: check if answer contains key words
        is_correct = student_ans == correct_ans or (len(correct_ans) > 3 and correct_ans in student_ans)
        evaluations.append({
            "question_id": qid,
            "is_correct": is_correct,
            "marks_awarded": q["marks"] if is_correct else 0,
            "ai_feedback": "Correct! 🎉" if is_correct else f"The answer is: {q['answer_key']}",
        })

    # Evaluate subjective questions with LLM
    if subjective_ids:
        subj_lines = []
        for qid in subjective_ids:
            q = q_map.get(qid)
            if not q:
                continue
            student_ans = answer_map.get(qid) or "(no answer)"
            subj_lines.append(
                f"Q: {q['question']}\nExpected: {q['answer_key']}\nStudent answered: {student_ans}\nMax marks: {q['marks']}"
            )

        sys_p = (
            f"You are evaluating quiz answers for {req.class_name} students.\n"
            f"Be fair and encouraging. Award partial marks for partially correct answers.\n"
            f"Return ONLY valid JSON."
        )
        llm_p = f"""Evaluate these student answers:

{chr(10).join(subj_lines)}

Return JSON:
{{
  "evaluations": [
    {{
      "question_id": "<id>",
      "marks_awarded": <number>,
      "is_correct": <true/false>,
      "ai_feedback": "brief encouraging feedback"
    }}
  ]
}}"""

        try:
            result, _ = await _call_llm(llm_p, sys_p)
            start = result.find('{')
            end = result.rfind('}') + 1
            if start >= 0 and end > start:
                data = json.loads(result[start:end])
                for ev in data.get("evaluations", []):
                    q = q_map.get(ev.get("question_id"))
                    if q:
                        marks = min(int(ev.get("marks_awarded", 0)), q["marks"])
                        evaluations.append({
                            "question_id": ev["question_id"],
                            "is_correct": ev.get("is_correct", marks == q["marks"]),
                            "marks_awarded": marks,
                            "ai_feedback": ev.get("ai_feedback", ""),
                        })
        except Exception as e:
            print(f"[evaluate-quiz] LLM error: {e}")
            # Fallback: 0 marks for subjective
            for qid in subjective_ids:
                q = q_map.get(qid)
                if q:
                    evaluations.append({
                        "question_id": qid,
                        "is_correct": False,
                        "marks_awarded": 0,
                        "ai_feedback": f"Expected: {q['answer_key']}",
                    })

    return {"evaluations": evaluations}


# ---------------------------------------------------------------------------
# TOC extraction
# ---------------------------------------------------------------------------

def _parse_toc_lines(lines: list[str]) -> list[dict]:
    """
    Parse cleaned TOC lines into chapter dicts using regex.
    Handles OCR artifacts (l/1 confusion), en-dash ranges, single pages.
    """
    import re

    def fix_ocr(s: str) -> str:
        """Fix common OCR artifacts in page numbers: lowercase l → 1, O → 0."""
        return s.replace('l', '1').replace('O', '0').replace('o', '0')

    # Page range at end of line: digits (or OCR'd digits) separated by dash/en-dash
    PAGE_RANGE_RE = re.compile(r'([0-9lLoO]+)\s*[-–—]\s*([0-9lLoO]+)\s*$')
    PAGE_SINGLE_RE = re.compile(r'\b([0-9lLoO]+)\s*$')
    SEP_RE = re.compile(r'[\s.·•_\-–—]{2,}')

    chapters = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if re.fullmatch(r'(table\s+of\s+)?contents?', line, re.IGNORECASE):
            continue

        m_range = PAGE_RANGE_RE.search(line)
        m_single = PAGE_SINGLE_RE.search(line)

        if m_range:
            try:
                page_start = int(fix_ocr(m_range.group(1)))
                page_end = int(fix_ocr(m_range.group(2)))
            except ValueError:
                continue
            title_raw = line[:m_range.start()]
        elif m_single:
            try:
                page_start = int(fix_ocr(m_single.group(1)))
                page_end = page_start
            except ValueError:
                continue
            title_raw = line[:m_single.start()]
        else:
            continue

        title = SEP_RE.sub(' ', title_raw).strip().strip('.-–—').strip()
        if not title:
            continue

        chapters.append({
            "title": title,
            "topics": [],
            "page_start": page_start,
            "page_end": page_end,
        })

    return chapters


def _parse_toc_two_column(words: list[dict]) -> list[dict]:
    """
    Handle TOC layouts where chapter titles and page numbers are in separate
    columns (like the image: title on left, decorative line in middle, page on right).

    Uses x-position to split words into left/right columns, pairs them by y-row.
    """
    import re

    if not words:
        return []

    # Find the x midpoint of the page to split left (titles) vs right (pages)
    x_positions = [w["x0"] for w in words]
    x_mid = (min(x_positions) + max(x_positions)) / 2

    # Group words by y-row (rounded)
    rows: dict[int, list] = {}
    for w in words:
        y_key = round(w["top"] / 4) * 4
        rows.setdefault(y_key, []).append(w)

    def fix_ocr(s: str) -> str:
        return s.replace('l', '1').replace('O', '0').replace('o', '0')

    PAGE_RANGE_RE = re.compile(r'^([0-9lLoO]+)\s*[-–—]\s*([0-9lLoO]+)$')
    PAGE_SINGLE_RE = re.compile(r'^([0-9lLoO]+)$')
    SEP_RE = re.compile(r'[\s.·•_\-–—]{2,}')

    chapters = []
    for y_key in sorted(rows.keys()):
        row_words = sorted(rows[y_key], key=lambda w: w["x0"])

        # Split into left-column words (titles) and right-column words (page numbers)
        # Filter out purely decorative tokens first
        meaningful = [w for w in row_words if not re.fullmatch(r'[-–—_.·•\s]+', w["text"])]
        if not meaningful:
            continue

        left_words = [w["text"] for w in meaningful if w["x0"] <= x_mid]
        right_words = [w["text"] for w in meaningful if w["x0"] > x_mid]

        title_raw = " ".join(left_words).strip()
        page_raw = " ".join(right_words).strip()

        if not title_raw:
            continue

        # Skip header
        if re.fullmatch(r'(table\s+of\s+)?contents?', title_raw, re.IGNORECASE):
            continue

        # Parse page from right column
        page_start = page_end = None
        if page_raw:
            # Remove spaces around dash for matching
            page_compact = re.sub(r'\s*[-–—]\s*', '–', page_raw)
            m_range = PAGE_RANGE_RE.match(page_compact)
            m_single = PAGE_SINGLE_RE.match(page_compact)
            if m_range:
                try:
                    page_start = int(fix_ocr(m_range.group(1)))
                    page_end = int(fix_ocr(m_range.group(2)))
                except ValueError:
                    pass
            elif m_single:
                try:
                    page_start = int(fix_ocr(m_single.group(1)))
                    page_end = page_start
                except ValueError:
                    pass

        # Clean title
        title = SEP_RE.sub(' ', title_raw).strip().strip('.-–—').strip()
        if not title:
            continue

        chapters.append({
            "title": title,
            "topics": [],
            "page_start": page_start,
            "page_end": page_end,
        })

    return chapters


def _ocr_pdf_page(pdf_path: str, page_index: int) -> list[dict]:
    """
    Render a PDF page to an image using PyMuPDF and run Tesseract OCR on it.
    Returns word dicts with keys: text, x0, top — same shape as pdfplumber's
    extract_words() so the same parsers can consume the output directly.

    Requires: pytesseract + Tesseract binary installed, Pillow, PyMuPDF.
    On Windows set TESSERACT_PATH in .env, e.g.:
      TESSERACT_PATH=C:\\Program Files\\Tesseract-OCR\\tesseract.exe
    """
    import os
    import fitz  # PyMuPDF
    import pytesseract
    from PIL import Image
    import io

    # Windows: pytesseract needs the explicit path to tesseract.exe
    tesseract_path = os.getenv("TESSERACT_PATH", "")
    if tesseract_path:
        pytesseract.pytesseract.tesseract_cmd = tesseract_path

    doc = fitz.open(pdf_path)
    page = doc[page_index]
    # Render at 2x scale for better OCR accuracy on small text
    mat = fitz.Matrix(2.0, 2.0)
    pix = page.get_pixmap(matrix=mat)
    img_bytes = pix.tobytes("png")
    doc.close()

    img = Image.open(io.BytesIO(img_bytes))
    data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT, config='--psm 6')

    words = []
    for i, text in enumerate(data['text']):
        text = text.strip()
        if not text or int(data['conf'][i]) < 30:  # skip low-confidence tokens
            continue
        # Scale coordinates back to PDF points (rendered at 2x)
        words.append({
            'text': text,
            'x0': data['left'][i] / 2.0,
            'top': data['top'][i] / 2.0,
        })

    return words


@app.post("/internal/extract-toc")
async def extract_toc(file: UploadFile = File(...), toc_page: int = Form(1)):
    """
    Extract chapter/topic structure from a PDF's Table of Contents page.

    Strategy (in order):
      1. Render the page to a PNG image and send to vision LLM (Gemini / GPT-4o).
         This works for ALL PDF types — text-based, scanned, decorative layouts.
      2. If vision LLM fails or is unavailable, fall back to pdfplumber word
         extraction + regex parsing (fast, free, no API cost).
      3. If regex finds < 2 entries, try text LLM with structured row data.

    Returns:
      { "chapters": [{ "title", "topics": [], "page_start", "page_end" }],
        "failed": bool, "page_count": int }
    """
    import tempfile, os, json, re, base64
    import fitz  # PyMuPDF — for rendering page to image
    from query_pipeline import _call_vision_llm, _call_llm

    content = await file.read()
    suffix = os.path.splitext(file.filename or "upload.pdf")[1].lower() or ".pdf"

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        # Get page count
        import pdfplumber
        with pdfplumber.open(tmp_path) as pdf:
            total_pages = len(pdf.pages)

        if toc_page < 1 or toc_page > total_pages:
            raise HTTPException(
                status_code=422,
                detail=f"toc_page {toc_page} is out of range (PDF has {total_pages} pages)",
            )

        # ── Strategy 1: Vision LLM (works for ALL PDF types) ─────────────────
        # Render page to PNG at 2x resolution and send to Gemini/GPT-4o vision.
        # This is the most reliable approach — same as what ChatGPT does.
        try:
            doc = fitz.open(tmp_path)
            pg = doc[toc_page - 1]
            pix = pg.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))  # 1.5x is sufficient for TOC text
            img_bytes = pix.tobytes("png")
            doc.close()
            image_b64 = base64.b64encode(img_bytes).decode()

            vision_system = (
                "You are a curriculum assistant reading a textbook Table of Contents page.\n"
                "Return ONLY valid JSON — no markdown, no explanation outside the JSON.\n"
                "\n"
                "Output format:\n"
                '{"chapters": [{"title": "string", "page_start": int_or_null, "page_end": int_or_null, '
                '"topics": [{"name": "string", "page_start": int_or_null}]}]}\n'
                "\n"
                "Rules:\n"
                "1. Extract EVERY unit/chapter/section visible in the TOC as a chapter entry.\n"
                "2. Sub-headings or lessons listed UNDER a chapter (indented or grouped) go into\n"
                "   that chapter's 'topics' array, each with their own page number if shown.\n"
                "   Example: 'Unit 4: Seasons' is the chapter; 'The Four Seasons 106' and\n"
                "   'Anandi's Rainbow 115' are topics with page_start 106 and 115.\n"
                "3. Page numbers MUST be integers. Look carefully at the right side of each row.\n"
                "   Ranges '4 - 11', '4–11' → page_start=4, page_end=11.\n"
                "   Single page '3' → page_start=3, page_end=3.\n"
                "   No page visible → null.\n"
                "4. Ignore decorative elements: dots, dashes, lines between title and page number.\n"
                "5. Titles with dashes like 'Letters A – D' or 'Unit 1 – Forces' are chapter names.\n"
                "6. If the page does not contain a Table of Contents, return {\"chapters\": []}."
            )
            vision_prompt = "Extract the complete Table of Contents from this textbook page."

            result, provider = await _call_vision_llm(image_b64, vision_prompt, vision_system)
            print(f"[extract-toc] vision LLM ({provider}): {result[:300]}")

            if result:
                s = result.find("{")
                e = result.rfind("}") + 1
                if s >= 0 and e > s:
                    data = json.loads(result[s:e])
                    chapters = data.get("chapters") or []
                    if chapters:
                        print(f"[extract-toc] vision extracted {len(chapters)} chapters")
                        return {"chapters": chapters, "failed": False, "page_count": total_pages}
                    else:
                        print("[extract-toc] vision returned empty chapters list")

        except Exception as vision_err:
            print(f"[extract-toc] vision LLM failed: {vision_err}")

        # ── Strategy 2: pdfplumber word extraction + regex ────────────────────
        # Fast, free, no API. Works well for clean text-based PDFs.
        print("[extract-toc] falling back to regex parser")
        try:
            with pdfplumber.open(tmp_path) as pdf:
                page = pdf.pages[toc_page - 1]
                words = page.extract_words(
                    x_tolerance=5, y_tolerance=5,
                    keep_blank_chars=False, use_text_flow=False,
                )

            if words:
                sample = [(w['text'], round(w['x0']), round(w['top'])) for w in words[:20]]
                print(f"[extract-toc] pdfplumber words sample: {sample}")

                chapters = _parse_toc_two_column(words)
                if len(chapters) >= 2:
                    print(f"[extract-toc] two-column regex found {len(chapters)} chapters")
                    return {"chapters": chapters, "failed": False, "page_count": total_pages}

                # Build cleaned lines and try line parser
                lines_map: dict[int, list] = {}
                for w in words:
                    y_key = round(w["top"] / 3) * 3
                    lines_map.setdefault(y_key, []).append(w)
                cleaned_lines = []
                for y_key in sorted(lines_map.keys()):
                    lw = sorted(lines_map[y_key], key=lambda w: w["x0"])
                    meaningful = [w["text"] for w in lw if not re.fullmatch(r'[-–—_.·•\s]+', w["text"])]
                    if meaningful:
                        cleaned_lines.append(" ".join(meaningful))

                chapters = _parse_toc_lines(cleaned_lines)
                if len(chapters) >= 2:
                    print(f"[extract-toc] line regex found {len(chapters)} chapters")
                    return {"chapters": chapters, "failed": False, "page_count": total_pages}

        except Exception as regex_err:
            print(f"[extract-toc] regex parser failed: {regex_err}")
            cleaned_lines = []

        # ── Strategy 3: text LLM with structured rows ─────────────────────────
        print("[extract-toc] falling back to text LLM")
        try:
            with pdfplumber.open(tmp_path) as pdf:
                raw_text = pdf.pages[toc_page - 1].extract_text() or ""
            page_text = "\n".join(
                re.sub(r'[-–—_.·•]{3,}', ' ', line).strip()
                for line in raw_text.splitlines()
                if re.sub(r'[-–—_.·•]{3,}', ' ', line).strip()
            )

            if page_text:
                system = (
                    "You are a curriculum assistant. Extract the Table of Contents from the text.\n"
                    "Return ONLY valid JSON: "
                    '{"chapters": [{"title": "string", "topics": [], "page_start": int_or_null, "page_end": int_or_null}]}\n'
                    "Page ranges like '4 - 11' → page_start=4, page_end=11. Single '3' → page_start=3, page_end=3.\n"
                    "Ignore decorative separators. Titles with dashes like 'Letters A – D' are chapter names."
                )
                result, _ = await _call_llm(page_text, system)
                if result:
                    s = result.find("{")
                    e = result.rfind("}") + 1
                    if s >= 0 and e > s:
                        data = json.loads(result[s:e])
                        chapters = data.get("chapters") or []
                        if chapters:
                            return {"chapters": chapters, "failed": False, "page_count": total_pages}
        except Exception as text_llm_err:
            print(f"[extract-toc] text LLM failed: {text_llm_err}")

        return {
            "chapters": [],
            "failed": True,
            "page_count": total_pages,
            "reason": "Could not extract chapters from this page. Try a different page or use the Excel import option.",
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"[extract-toc] Error: {e}")
        return {"chapters": [], "failed": True, "page_count": 0}
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Textbook planner generation
# ---------------------------------------------------------------------------

class TextbookPlannerRequest(BaseModel):
    subjects: list = []          # list of subject dicts with chapters, weekly_hours, etc.
    parameters: dict = {}        # school parameters (timings, breaks, activities)
    test_config: dict = {}       # test scheduling config
    academic_year: str = ""
    working_days: list = []      # isoweekday ints [1..7]
    holidays: list = []          # ISO date strings
    special_days: list = []      # list of {day_date, day_type, duration_type}
    start_date: str = ""         # academic year start (ISO)
    end_date: str = ""           # academic year end (ISO)
    preview_only: bool = False   # if True, generate only up to preview_end_date
    preview_end_date: str = ""   # ISO date — limit entries to this date when preview_only=True


@app.post("/internal/generate-textbook-planner")
async def generate_textbook_planner(req: TextbookPlannerRequest):
    """
    Generate a day-by-day textbook planner from session configuration.

    Returns:
      {
        "entries": [{ date, subject_id, subject_name, chapter_name, topic_name, duration_minutes }],
        "summary": { total_teaching_days, total_exam_days, total_revision_days,
                     subjects: [{ name, coverage_pct }] }
      }
    """
    from planner_engine import (
        calculate_chapter_weights,
        calculate_available_minutes,
        get_teaching_days,
        insert_test_days,
        insert_revision_buffers,
        distribute_topics_with_llm,
        distribute_topics_across_days,
    )

    # Parse dates
    try:
        start_date = date.fromisoformat(req.start_date)
        end_date = date.fromisoformat(req.end_date)
    except (ValueError, TypeError) as e:
        raise HTTPException(status_code=422, detail=f"Invalid start_date or end_date: {e}")

    # Parse holidays
    holidays: list[date] = []
    for h in req.holidays:
        try:
            holidays.append(date.fromisoformat(str(h)[:10]))
        except ValueError:
            pass

    # Parse special_days
    special_days: list[dict] = []
    for sd in req.special_days:
        try:
            d = date.fromisoformat(str(sd.get("day_date", ""))[:10])
            special_days.append({
                "day_date": d,
                "day_type": sd.get("day_type", "event"),
                "duration_type": sd.get("duration_type", "full_day"),
            })
        except ValueError:
            pass

    working_days = [int(w) for w in req.working_days] if req.working_days else [1, 2, 3, 4, 5]

    # Calculate available minutes (informational; don't block generation on error)
    try:
        if req.parameters:
            calculate_available_minutes(req.parameters)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    # Step 1: Get initial teaching days (before exam/revision days are known)
    initial_teaching_days = get_teaching_days(
        start_date, end_date, working_days, holidays, special_days
    )

    # Step 2: Compute chapter weights per subject
    subjects_with_weights = []
    for subj in req.subjects:
        chapters = calculate_chapter_weights(subj.get("chapters") or [])
        subjects_with_weights.append({**subj, "chapters": chapters})

    # Step 3: Insert test days
    test_config = req.test_config or {}
    mode = test_config.get("mode", "manual")

    # Build chapters_by_subject for end-of-chapter mode
    # (requires last_teaching_day per chapter — computed after distribution)
    chapters_by_subject: dict[str, list[dict]] = {}
    for subj in subjects_with_weights:
        chapters_by_subject[subj.get("subject_name", "")] = subj.get("chapters") or []

    exam_days, _ = insert_test_days(mode, test_config, initial_teaching_days, chapters_by_subject)

    # Step 4: Insert revision buffers
    revision_days: list[date] = []
    if test_config.get("revision_buffer", True) and exam_days:
        revision_days = insert_revision_buffers(exam_days, initial_teaching_days)

    # Step 5: Recompute teaching days excluding exam and revision days
    exam_revision_special = special_days + [
        {"day_date": d, "day_type": "exam", "duration_type": "full_day"} for d in exam_days
    ] + [
        {"day_date": d, "day_type": "revision", "duration_type": "full_day"} for d in revision_days
    ]

    teaching_days = get_teaching_days(
        start_date, end_date, working_days, holidays, exam_revision_special
    )

    # Step 6: Distribute topics across teaching days using LLM for smart interleaving
    entries = await distribute_topics_with_llm(
        subjects_with_weights,
        teaching_days,
        exam_days,
        revision_days,
        holidays,
        req.parameters,
    )

    # Step 7: Build summary
    total_topics_by_subject: dict[str, int] = {}
    covered_topics_by_subject: dict[str, int] = {}
    for subj in subjects_with_weights:
        name = subj.get("subject_name", "")
        total = sum(len(ch.get("topics") or []) for ch in subj.get("chapters") or [])
        total_topics_by_subject[name] = total

    for entry in entries:
        name = entry.get("subject_name", "")
        covered_topics_by_subject[name] = covered_topics_by_subject.get(name, 0) + 1

    subjects_summary = []
    for subj in subjects_with_weights:
        name = subj.get("subject_name", "")
        total = total_topics_by_subject.get(name, 0)
        covered = covered_topics_by_subject.get(name, 0)
        pct = round(covered / total * 100, 1) if total > 0 else 100.0
        subjects_summary.append({"name": name, "coverage_pct": pct})

    # If preview_only, filter entries to the preview window only
    if req.preview_only and req.preview_end_date:
        try:
            preview_cutoff = date.fromisoformat(req.preview_end_date)
            entries = [e for e in entries if date.fromisoformat(e["date"]) <= preview_cutoff]
        except ValueError:
            pass  # invalid date — return all entries

    return {
        "entries": entries,
        "summary": {
            "total_teaching_days": len(teaching_days),
            "total_exam_days": len(exam_days),
            "total_revision_days": len(revision_days),
            "subjects": subjects_summary,
            "preview_only": req.preview_only,
        },
    }
