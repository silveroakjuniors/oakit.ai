from fastapi import FastAPI, BackgroundTasks, UploadFile, File, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from dotenv import load_dotenv
from datetime import datetime, date
import hashlib
import io

load_dotenv()

app = FastAPI(title="Oakit AI Service", version="0.1.0")

THOUGHTS_FOR_DAY = [
    "A great teacher doesn't just teach — they inspire curiosity.",
    "Every child learns differently. Patience is your greatest tool.",
    "Small moments of encouragement can change a student's entire day.",
    "The best classroom is one where every student feels seen.",
    "Teaching is the one profession that creates all other professions.",
    "Your energy sets the tone for the entire classroom.",
    "A smile and a kind word can unlock a child's potential.",
    "Celebrate small wins — they build the confidence for big ones.",
    "The questions children ask are the seeds of tomorrow's discoveries.",
    "You are not just teaching a subject — you are shaping a future.",
    "Consistency and care are the foundations of great teaching.",
    "Every lesson is a chance to make learning feel like an adventure.",
    "The most powerful thing you can give a student is belief in themselves.",
    "Great teachers listen as much as they speak.",
    "Your dedication today plants seeds that bloom for a lifetime.",
]


@app.get("/health")
async def health():
    return {"status": "ok", "service": "oakit-ai-service"}


# --- Plan-scoped questions generator ---

class PlanQuestionsRequest(BaseModel):
    teacher_id: str
    school_id: str
    query_date: str

@app.post("/internal/plan-questions")
async def plan_questions(req: PlanQuestionsRequest):
    """Generate suggested questions based on today's plan — strictly scoped to curriculum."""
    from query_pipeline import generate_plan_questions
    result = await generate_plan_questions(req.teacher_id, req.school_id, req.query_date)
    return result


# --- Curriculum preview (Week 1 only, before full ingestion) ---

@app.post("/internal/preview")
async def preview_curriculum(file: UploadFile = File(...), start_page: str = "1"):
    """Extract only Week 1 from the PDF and return it for admin approval."""
    import tempfile, os
    from extractor import extract_days, extract_pages
    from chunker import chunk_from_days

    # start_page comes as form string
    try:
        start_page_int = int(start_page)
    except (ValueError, TypeError):
        start_page_int = 1

    content = await file.read()
    suffix = os.path.splitext(file.filename or "upload.pdf")[1] or ".pdf"

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        day_entries, failed_pages = extract_days(tmp_path, start_page=start_page_int)
        # Filter to Week 1 only
        week1 = [d for d in day_entries if d.week == 1]

        if not week1:
            # Fallback: return first 5 text chunks
            pages, _ = extract_pages(tmp_path, start_page=start_page_int)
            from chunker import chunk_document
            chunks = chunk_document(pages)[:5]
            return {
                "mode": "text",
                "week": 1,
                "total_weeks_detected": 0,
                "total_days_detected": len(day_entries),
                "days": [{"week": 0, "day": i+1, "topic_label": c.topic_label, "subjects": {"Content": c.content[:500]}} for i, c in enumerate(chunks)],
                "failed_pages": failed_pages,
            }

        total_weeks = max((d.week for d in day_entries), default=0)
        total_days = len(day_entries)

        return {
            "mode": "table",
            "week": 1,
            "total_weeks_detected": total_weeks,
            "total_days_detected": total_days,
            "days": [
                {
                    "week": d.week,
                    "day": d.day,
                    "topic_label": d.topic_label,
                    "subjects": d.subjects,
                }
                for d in week1
            ],
            "failed_pages": failed_pages,
        }
    except Exception as e:
        import traceback
        print(f"Preview error: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


# --- Ingestion ---

class IngestRequest(BaseModel):
    document_id: str

@app.post("/internal/ingest")
async def ingest(req: IngestRequest, background_tasks: BackgroundTasks):
    background_tasks.add_task(_run_ingest, req.document_id)
    return {"message": "Ingestion started", "document_id": req.document_id}

async def _run_ingest(document_id: str):
    try:
        from ingestion_service import ingest_document
        result = await ingest_document(document_id)
        print(f"Ingestion complete for {document_id}: {result}")
    except Exception as e:
        print(f"Ingestion failed for {document_id}: {e}")


# --- Query ---

class QueryRequest(BaseModel):
    teacher_id: str
    school_id: str
    text: str
    query_date: str
    role: str = "teacher"
    history: list = []

@app.post("/internal/query")
async def query_endpoint(req: QueryRequest):
    from query_pipeline import query
    result = await query(req.teacher_id, req.school_id, req.text, req.query_date, req.role, history=req.history)
    return result


# --- Plan generation ---

class PlanRequest(BaseModel):
    class_id: str
    section_id: str
    school_id: str
    academic_year: str
    month: int = None
    plan_year: int = None

@app.post("/internal/generate-plans")
async def generate_plans(req: PlanRequest):
    from planner_service import generate_plans
    count = await generate_plans(
        req.class_id, req.section_id, req.school_id, req.academic_year,
        month=req.month, plan_year=req.plan_year
    )
    return {"plans_created": count}


# --- Coverage analysis ---

class CoverageRequest(BaseModel):
    coverage_log_id: str
    log_text: str
    section_id: str
    log_date: str

@app.post("/internal/analyze-coverage")
async def analyze_coverage(req: CoverageRequest):
    from coverage_analyzer import analyze_coverage
    results = await analyze_coverage(req.coverage_log_id, req.log_text, req.section_id, req.log_date)
    return {"results": results}


# --- Holiday import ---

@app.post("/internal/import-holidays")
async def import_holidays(file: UploadFile = File(...)):
    import openpyxl
    content = await file.read()
    valid_rows = []
    invalid_rows = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

        # Flexible column detection — support multiple naming conventions
        def find_col(candidates):
            for c in candidates:
                if c in headers:
                    return headers.index(c)
            return None

        date_idx = find_col(['date'])
        # event_name can come from 'description', 'event_name', 'event', 'name', 'holiday name'
        name_idx = find_col(['description', 'event_name', 'event', 'name', 'holiday name', 'holiday'])
        # type column: 'type', 'day type', 'category'
        type_idx = find_col(['type', 'day type', 'category'])

        if date_idx is None:
            raise HTTPException(status_code=400, detail="Missing required column: Date")
        if name_idx is None:
            raise HTTPException(status_code=400, detail="Missing required column: Description or event_name")

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_date = row[date_idx] if date_idx < len(row) else None
            event_name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
            row_type = str(row[type_idx]).strip().lower() if type_idx is not None and type_idx < len(row) and row[type_idx] else ''

            if not event_name or not str(event_name).strip():
                invalid_rows.append({"row": row_num, "reason": "Missing description/event name"})
                continue

            # Skip rows that are explicitly "Working Day" — those aren't holidays
            if row_type in ('working day', 'working', 'school day'):
                invalid_rows.append({"row": row_num, "reason": f"Skipped: type is '{row_type}' (not a holiday)"})
                continue

            # Parse date
            parsed_date = None
            if isinstance(raw_date, (datetime, date)):
                parsed_date = raw_date.strftime('%Y-%m-%d') if isinstance(raw_date, datetime) else str(raw_date)
            elif isinstance(raw_date, str):
                raw_str = raw_date.strip()
                for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%d %b %Y', '%d %B %Y', '%B %d %Y'):
                    try:
                        parsed_date = datetime.strptime(raw_str, fmt).strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        continue

            if not parsed_date:
                invalid_rows.append({"row": row_num, "reason": f"Invalid date: {raw_date}"})
                continue

            valid_rows.append({
                "date": parsed_date,
                "event_name": str(event_name).strip(),
                "type": row_type or "holiday"
            })

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    return {"valid_rows": valid_rows, "invalid_rows": invalid_rows}


# --- Student import ---

REQUIRED_STUDENT_COLUMNS = {'student name', 'father name', 'section', 'class', 'parent contact number'}

@app.post("/internal/import-students")
async def import_students(file: UploadFile = File(...)):
    import openpyxl
    content = await file.read()
    valid_rows = []
    invalid_rows = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

        missing = REQUIRED_STUDENT_COLUMNS - set(headers)
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(sorted(missing))}")

        col = {h: i for i, h in enumerate(headers)}

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            def get(key):
                idx = col.get(key)
                return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ''

            student_name = get('student name')
            father_name = get('father name')
            section = get('section')
            class_name = get('class')
            parent_contact = get('parent contact number')

            if not student_name:
                invalid_rows.append({"row": row_num, "reason": "Missing student name"})
                continue

            valid_rows.append({
                "student_name": student_name,
                "father_name": father_name,
                "section": section,
                "class": class_name,
                "parent_contact": parent_contact,
            })

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    return {"valid_rows": valid_rows, "invalid_rows": invalid_rows}


# --- Holiday list PDF export ---

class HolidayExportRequest(BaseModel):
    academic_year: str
    holidays: list

@app.post("/internal/export-holiday-pdf")
async def export_holiday_pdf(req: HolidayExportRequest):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.colors import Color
    from datetime import date as date_type
    import io

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    DAY_NAMES = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']

    def draw_watermark(canvas):
        canvas.saveState()
        canvas.setFont("Helvetica", 60)
        canvas.setFillColor(Color(0.85, 0.85, 0.85, alpha=0.3))
        canvas.translate(width / 2, height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "OAKIT.AI")
        canvas.restoreState()

    draw_watermark(c)
    y = height - 2 * cm

    c.setFont("Helvetica-Bold", 16)
    c.drawString(2 * cm, y, f"Holiday List — {req.academic_year}")
    y -= 0.6 * cm
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    from datetime import datetime
    c.drawString(2 * cm, y, f"Generated on {datetime.now().strftime('%d %B %Y')}   ·   {len(req.holidays)} holidays")
    c.setFillColorRGB(0, 0, 0)
    y -= 0.4 * cm
    c.line(2 * cm, y, width - 2 * cm, y)
    y -= 0.7 * cm

    # Table header
    c.setFont("Helvetica-Bold", 9)
    c.setFillColorRGB(0.3, 0.3, 0.3)
    c.drawString(2 * cm, y, "#")
    c.drawString(2.8 * cm, y, "Holiday")
    c.drawString(12 * cm, y, "Day")
    c.drawString(14.5 * cm, y, "Date")
    c.setFillColorRGB(0, 0, 0)
    y -= 0.3 * cm
    c.line(2 * cm, y, width - 2 * cm, y)
    y -= 0.5 * cm

    for i, h in enumerate(req.holidays, 1):
        if y < 3 * cm:
            c.showPage()
            draw_watermark(c)
            y = height - 2 * cm

        try:
            d = date_type.fromisoformat(str(h.get('date', ''))[:10])
            day_name = DAY_NAMES[d.weekday()]
            date_str = d.strftime('%d %b %Y')
        except Exception:
            day_name = ''
            date_str = str(h.get('date', ''))[:10]

        # Alternating row background
        if i % 2 == 0:
            c.setFillColorRGB(0.97, 0.97, 0.97)
            c.rect(1.8 * cm, y - 0.15 * cm, width - 3.6 * cm, 0.55 * cm, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)

        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0.5, 0.5, 0.5)
        c.drawString(2 * cm, y, str(i))
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(2.8 * cm, y, h.get('event_name', '')[:55])
        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0.2, 0.5, 0.3)
        c.drawString(12 * cm, y, day_name)
        c.setFillColorRGB(0.3, 0.3, 0.3)
        c.drawString(14.5 * cm, y, date_str)
        c.setFillColorRGB(0, 0, 0)
        y -= 0.55 * cm

    c.save()
    buf.seek(0)
    from fastapi.responses import Response as FastResponse
    return FastResponse(content=buf.read(), media_type="application/pdf")


# --- Monthly plan PDF export ---

class MonthlyExportRequest(BaseModel):
    section_id: str
    section_label: str
    class_name: str
    month_name: str
    plans: list = []
    days: list = []

@app.post("/internal/export-monthly-pdf")
async def export_monthly_pdf(req: MonthlyExportRequest):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.colors import Color
    import io

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    def draw_watermark(canvas):
        canvas.saveState()
        canvas.setFont("Helvetica", 60)
        canvas.setFillColor(Color(0.85, 0.85, 0.85, alpha=0.3))
        canvas.translate(width / 2, height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "OAKIT.AI")
        canvas.restoreState()

    draw_watermark(c)
    y = height - 2 * cm

    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(2 * cm, y, f"Monthly Plan — {req.month_name}")
    y -= 0.7 * cm
    c.setFont("Helvetica", 11)
    c.drawString(2 * cm, y, f"Class: {req.class_name}   |   Section: {req.section_label}")
    y -= 0.5 * cm
    c.line(2 * cm, y, width - 2 * cm, y)
    y -= 0.7 * cm

    for plan in (req.plans or []):
        if y < 4 * cm:
            c.showPage()
            draw_watermark(c)
            y = height - 2 * cm

        day_type = plan.get('type', 'working')
        plan_date = plan.get('date') or plan.get('plan_date', '')
        try:
            from datetime import date as date_type
            d = date_type.fromisoformat(str(plan_date)[:10])
            date_str = d.strftime('%a, %d %b')
        except Exception:
            date_str = str(plan_date)[:10]

        if day_type == 'weekend':
            c.setFont("Helvetica", 9)
            c.setFillColorRGB(0.7, 0.7, 0.7)
            c.drawString(2 * cm, y, f"{date_str}   —   Weekend")
            c.setFillColorRGB(0, 0, 0)
            y -= 0.45 * cm

        elif day_type == 'holiday':
            c.setFont("Helvetica-Bold", 9)
            c.setFillColorRGB(0.8, 0.1, 0.1)
            holiday_label = plan.get('label', 'Holiday')
            c.drawString(2 * cm, y, f"{date_str}   \U0001f389  {holiday_label}")
            c.setFillColorRGB(0, 0, 0)
            y -= 0.45 * cm

        else:
            c.setFont("Helvetica-Bold", 10)
            c.setFillColorRGB(0.1, 0.3, 0.2)
            c.drawString(2 * cm, y, date_str)
            c.setFillColorRGB(0, 0, 0)
            y -= 0.45 * cm

            chunks = plan.get('chunks', [])
            if not chunks:
                c.setFont("Helvetica-Oblique", 8)
                c.setFillColorRGB(0.6, 0.6, 0.6)
                c.drawString(2.5 * cm, y, "No plan assigned")
                c.setFillColorRGB(0, 0, 0)
                y -= 0.4 * cm
            else:
                for chunk in chunks:
                    if y < 3 * cm:
                        c.showPage()
                        draw_watermark(c)
                        y = height - 2 * cm
                    c.setFont("Helvetica-Bold", 9)
                    topic = chunk.get('topic_label', 'Topic')
                    c.drawString(2.5 * cm, y, f"• {topic}")
                    y -= 0.4 * cm

                    content = chunk.get('content', '').strip()
                    if content:
                        for line in content.split('\n'):
                            line = line.strip()
                            if not line:
                                continue
                            if y < 3 * cm:
                                c.showPage()
                                draw_watermark(c)
                                y = height - 2 * cm
                            if ':' in line:
                                parts = line.split(':', 1)
                                subject = parts[0].strip()
                                activity = parts[1].strip()
                                c.setFont("Helvetica-Bold", 8)
                                c.setFillColorRGB(0.3, 0.3, 0.3)
                                c.drawString(3 * cm, y, f"{subject}:")
                                c.setFont("Helvetica", 8)
                                c.setFillColorRGB(0, 0, 0)
                                max_chars = 80
                                if len(activity) > max_chars:
                                    c.drawString(5.5 * cm, y, activity[:max_chars])
                                    y -= 0.35 * cm
                                    c.drawString(5.5 * cm, y, activity[max_chars:max_chars*2])
                                else:
                                    c.drawString(5.5 * cm, y, activity)
                            else:
                                c.setFont("Helvetica", 8)
                                c.drawString(3 * cm, y, line[:100])
                            y -= 0.35 * cm

                    activity_ids = chunk.get('activity_ids', [])
                    if activity_ids:
                        if y < 3 * cm:
                            c.showPage()
                            draw_watermark(c)
                            y = height - 2 * cm
                        c.setFont("Helvetica-Oblique", 7.5)
                        c.setFillColorRGB(0.2, 0.5, 0.3)
                        c.drawString(3 * cm, y, f"\U0001f4ce {', '.join(activity_ids)}")
                        c.setFillColorRGB(0, 0, 0)
                        y -= 0.35 * cm

                    y -= 0.15 * cm

        y -= 0.1 * cm

    c.save()
    buf.seek(0)
    from fastapi.responses import Response as FastResponse
    return FastResponse(content=buf.read(), media_type="application/pdf")


# --- PDF export ---

class ExportPdfRequest(BaseModel):
    teacher_id: str
    section_id: str
    section_label: str = ""
    teacher_name: str = "Teacher"
    date: str
    days: int = 1
    settling_text: str = None

@app.post("/internal/export-pdf")
async def export_pdf(req: ExportPdfRequest):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.colors import Color
    from db import get_pool
    from uuid import UUID
    from datetime import date as date_type, timedelta

    pool = await get_pool()
    start_date = date_type.fromisoformat(req.date)

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    def draw_watermark(canvas):
        canvas.saveState()
        canvas.setFont("Helvetica", 60)
        canvas.setFillColor(Color(0.85, 0.85, 0.85, alpha=0.3))
        canvas.translate(width / 2, height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "OAKIT.AI")
        canvas.restoreState()

    # If settling_text provided — render the AI-generated plan as PDF
    if req.settling_text:
        draw_watermark(c)
        y = height - 2 * cm
        c.setFont("Helvetica-Bold", 14)
        c.drawString(2 * cm, y, f"Settling Day Plan — {start_date.strftime('%d %B %Y')}")
        y -= 0.7 * cm
        c.setFont("Helvetica", 11)
        c.drawString(2 * cm, y, f"Teacher: {req.teacher_name}   |   Section: {req.section_label}")
        y -= 0.5 * cm
        c.line(2 * cm, y, width - 2 * cm, y)
        y -= 0.7 * cm

        # Render each line of the settling text
        import re
        for line in req.settling_text.split('\n'):
            line = line.strip()
            if not line:
                y -= 0.3 * cm
                continue
            if y < 3 * cm:
                c.showPage()
                draw_watermark(c)
                y = height - 2 * cm
            # Strip emoji for PDF (reportlab doesn't support all emoji)
            clean = re.sub(r'[^\x00-\x7F]+', '', line).strip()
            if not clean:
                continue
            if line.startswith('📌') or line.startswith('What to do') or line.startswith('Tip'):
                c.setFont("Helvetica-Bold", 10)
            else:
                c.setFont("Helvetica", 9)
            c.drawString(2 * cm, y, clean[:100])
            y -= 0.45 * cm

        c.save()
        buf.seek(0)
        return Response(content=buf.read(), media_type="application/pdf")

    # Regular curriculum day export
    plan = await pool.fetchrow(
        "SELECT plan_date, chunk_ids FROM day_plans WHERE section_id = $1 AND plan_date = $2",
        UUID(req.section_id), start_date
    )
    if not plan or not plan['chunk_ids']:
        raise HTTPException(status_code=404, detail="No day plan found for the requested date")

    chunks = await pool.fetch(
        "SELECT topic_label, content, activity_ids FROM curriculum_chunks WHERE id = ANY($1::uuid[]) ORDER BY chunk_index",
        plan['chunk_ids']
    )

    draw_watermark(c)
    y = height - 2 * cm

    c.setFont("Helvetica-Bold", 14)
    c.drawString(2 * cm, y, f"Day Plan — {start_date.strftime('%d %B %Y')}")
    y -= 0.7 * cm
    c.setFont("Helvetica", 11)
    c.drawString(2 * cm, y, f"Teacher: {req.teacher_name}   |   Section: {req.section_label}")
    y -= 0.5 * cm
    c.line(2 * cm, y, width - 2 * cm, y)
    y -= 0.7 * cm

    for chunk in chunks:
        if y < 3 * cm:
            c.showPage()
            draw_watermark(c)
            y = height - 2 * cm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(2 * cm, y, f"• {chunk['topic_label'] or 'Topic'}")
        y -= 0.5 * cm

        content = (chunk['content'] or '').strip()
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue
            if y < 3 * cm:
                c.showPage()
                draw_watermark(c)
                y = height - 2 * cm
            c.setFont("Helvetica", 9)
            c.drawString(2.5 * cm, y, line[:100])
            y -= 0.4 * cm

        if chunk['activity_ids']:
            c.setFont("Helvetica-Oblique", 9)
            c.setFillColorRGB(0.2, 0.5, 0.3)
            c.drawString(2.5 * cm, y, f"Materials: {', '.join(chunk['activity_ids'])}")
            c.setFillColorRGB(0, 0, 0)
            y -= 0.4 * cm

        y -= 0.3 * cm

    c.save()
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf")


# --- Greeting ---

@app.get("/internal/greeting")
async def greeting(teacher_name: str = "Teacher", teacher_id: str = ""):
    now = datetime.now()
    hour = now.hour

    if 5 <= hour < 12:
        salutation = "Good morning"
    elif 12 <= hour < 17:
        salutation = "Good afternoon"
    else:
        salutation = "Good evening"

    greeting_text = f"{salutation}, {teacher_name}!"

    # Early arrival (before 7am)
    if hour < 7:
        greeting_text += " You're here early — that's dedication!"

    # Rotate thought by date hash
    date_str = now.strftime('%Y-%m-%d') + teacher_id
    idx = int(hashlib.md5(date_str.encode()).hexdigest(), 16) % len(THOUGHTS_FOR_DAY)
    thought = THOUGHTS_FOR_DAY[idx]

    return {
        "greeting": greeting_text,
        "thought_for_day": thought,
    }
